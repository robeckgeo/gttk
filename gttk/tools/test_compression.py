#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# ******************************************************************************
# Project: GeoTIFF ToolKit (GTTK)
# Author: Eric Robeck <robeckgeo@gmail.com>
#
# Copyright (c) 2025, Eric Robeck
# Licensed under the MIT License
# ******************************************************************************

"""
GeoTIFF Compression Testing and Benchmarking Tool for GTTK.

This module contains the logic for the 'test' command. It automates the process
of running multiple compression scenarios on a source GeoTIFF, measuring key
performance metrics like file size, write speed, and read speed, and generates
a summary report in Excel format.

Classes:
    ExcelWriter: Handles writing formatted results to an Excel template.
    TestResultMetrics: A dataclass for holding performance metrics.
    GeoTiffMetadata: A dataclass for storing extracted GeoTIFF metadata.
"""
import csv
import gc
import logging
import os
import sys
import time
import traceback
from copy import copy
from dataclasses import dataclass
from importlib import resources
from osgeo import gdal
from pathlib import Path
from typing import cast, Any, Dict, List, Optional, Tuple
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from gttk.main import str2bool
from gttk.tools.optimize_compression import optimize_compression
import gttk.tools.optimize_compression_arc as optimize_compression_arc
from gttk.utils.exceptions import CSVLoadError, OptimizationError
from gttk.utils.geotiff_processor import calculate_compression_efficiency, determine_decimal_precision
from gttk.utils.log_helpers import shutdown_logger, init_arcpy, ArcpyLogHandler
from gttk.utils.optimize_constants import CompressionAlgorithm as CA, ProductType as PT
from gttk.utils.path_helpers import get_geotiff_files
from gttk.utils.script_arguments import TestArguments, OptimizeArguments
from gttk.utils.validate_cloud_optimized_geotiff import validate as validate_cog

# --- Configuration & Setup ---
gdal.SetConfigOption('GDAL_NUM_THREADS', 'ALL_CPUS')
logger = logging.getLogger('test_compression')

# --- Constants ---
# Script Behavior
DEFAULT_TEMP_DIR = Path("./temp")
NUM_COMPRESSION_RUNS = 1
NUM_WARMUP_READ_ITERATIONS = 3
NUM_MEASURED_READ_ITERATIONS = 10

# GeoTIFF Types
TYPE_DEM: str = PT.DEM.value
TYPE_IMAGE: str = PT.IMAGE.value
TYPE_ERROR: str = PT.ERROR.value
TYPE_SCIENTIFIC: str = PT.SCIENTIFIC.value
TYPE_THEMATIC: str = PT.THEMATIC.value
TYPE_ORIGINAL: str = 'original'

# Compression Algorithms
ALGO_NONE: str = CA.NONE.value
ALGO_DEFLATE: str = CA.DEFLATE.value
ALGO_ZSTD: str = CA.ZSTD.value
ALGO_LERC: str = CA.LERC.value
ALGO_JPEG: str = CA.JPEG.value
ALGO_LZW: str = CA.LZW.value
ALGO_JXL: str = CA.JXL.value

# Predictor Values
PREDICTOR_STANDARD = '1'
PREDICTOR_HORIZONTAL = '2'
PREDICTOR_FLOATING_POINT = '3'

# Output Column Names
COL_FILENAME = 'filename'
COL_TYPE = 'type'
COL_COG = 'cog'
COL_OVERVIEWS = 'overviews'
COL_ALGORITHM = 'algorithm'
COL_TILE_SIZE = 'tile_size'
COL_LEVEL = 'level'
COL_PREDICTOR = 'predictor'
COL_DECIMALS = 'decimals'
COL_MAX_Z_ERROR = 'max_z_error'
COL_QUALITY = 'quality'
COL_NODATA = 'nodata'
COL_VERTICAL_SRS = 'vertical_srs'
COL_SIZE_MB = 'size (MB)'
COL_SIZE_DELTA = 'size_delta (%)'
COL_COMPRESSION_PERCENT = 'compression (%)'
COL_IMPROVEMENT_PERCENT = 'improvement (%)'
COL_WRITE_SPEED = 'write (Mb/CPU-sec)'
COL_WRITE_RATIO = 'write (ratio)'
COL_READ_SPEED = 'read (Mb/CPU-sec)'
COL_READ_RATIO = 'read (ratio)'
COL_COMMENT = 'comment'

OUTPUT_COLUMNS = [
    COL_FILENAME, COL_TYPE, COL_COG, COL_OVERVIEWS, COL_TILE_SIZE, COL_ALGORITHM, COL_LEVEL, COL_PREDICTOR,
    COL_DECIMALS, COL_MAX_Z_ERROR, COL_QUALITY, COL_SIZE_MB, COL_SIZE_DELTA, COL_COMPRESSION_PERCENT,
    COL_IMPROVEMENT_PERCENT, COL_WRITE_SPEED, COL_WRITE_RATIO, COL_READ_SPEED, COL_READ_RATIO, COL_COMMENT
]

csv_file_MANDATORY_HEADERS = [COL_TYPE, COL_COG, COL_OVERVIEWS, COL_ALGORITHM]

# Default/Error Values
VAL_NA = ''
VAL_ERROR = 'Error'
VAL_SKIPPED = ''

# --- Helper Functions ---
def _get_optimize_script_path(arc_mode: bool = False) -> Path:
    """Select appropriate optimization script based on execution mode."""
    script_dir = Path(__file__).parent.resolve()
    if arc_mode:
        return script_dir / "optimize_compression_arc.py"
    else:
        return script_dir / "optimize_compression.py"

def _load_compression_options(product_type: str) -> List[Dict[str, str]]:
    """Load appropriate CSV based on data type."""
    # Templates are in gttk/resources/templates
    csv_file = Path(__file__).parent.parent / "resources" / "templates" / f"compression_options_{product_type}.csv"
    
    if not csv_file.exists():
        available_types = [TYPE_DEM, TYPE_IMAGE, TYPE_ERROR, TYPE_SCIENTIFIC, TYPE_THEMATIC]
        raise CSVLoadError(
            f"No compression options found for type: {product_type}. "
            f"Available types: {', '.join(available_types)}"
        )
    
    return _read_test_cases_from_csv(csv_file)

def _is_cog(filepath: Path) -> bool:
    """Checks if a GeoTIFF is a valid Cloud-Optimized GeoTIFF."""
    if not filepath.exists():
        return False
    try:
        warnings, errors, _ = validate_cog(str(filepath))
        return not errors
    except Exception as e:
        logger.warning(f"COG validation for {filepath.name} failed: {e}")
        return False

def _get_tile_size(ds: gdal.Dataset) -> Optional[int]:
    """Return True if dataset is tiled (not striped)."""
    if ds is None:
        return None
    try:
        band = ds.GetRasterBand(1)
        bx, by = band.GetBlockSize()
        
        # Strongest check: If block width is smaller than raster width, it implies tiling.
        # (Strips always span the full width of the raster).
        if bx < ds.RasterXSize:
            return bx
            
        # If block width equals raster width, it could be a Strip or a Single Tile.
        # Check metadata to resolve ambiguity.
        tiled_meta = ds.GetMetadataItem('TILED', 'IMAGE_STRUCTURE')
        if tiled_meta is not None and tiled_meta.upper() in ('YES', 'TRUE', '1'):
            return bx
            
        # Fallback heuristic: If block is square (e.g. 512x512) and covers full width, treat as tile.
        if bx == by:
            return bx
            
        return None # Assume striped
    finally:
        band = None

def _has_internal_overviews(filepath: Path) -> bool:
    """Checks if a GeoTIFF has internal overviews."""
    if not filepath.exists():
        return False

    try:
        ds = gdal.Open(str(filepath), gdal.GA_ReadOnly)
        if ds is None:
            logger.warning(f"Could not open {filepath.name} to check for overviews.")
            return False
        
        band = ds.GetRasterBand(1)
        count = band.GetOverviewCount() > 0
        if count == 0:
            return False

        # Consider internal only if every overview is stored in the TIFF itself
        for i in range(count):
            ovr_band = band.GetOverview(i)
            ovr_ds = ovr_band.GetDataset()
            ovr_files = ovr_ds.GetFileList() or []
            if ovr_files:  # external .ovr file(s) exist
                return False

        return True

    except Exception as e:
        logger.warning(f"Overview check for {filepath.name} failed: {e}")
        return False
    finally:
        # Break reference cycles by clearing band variables first
        ovr_band = None
        band = None
        ds = None

def _generate_temp_filename(source_name_stem: str, type_val: str, params_dict: Dict[str, Any], temp_dir_path: Path) -> Path:
    """Generates a unique temporary filename based on parameters."""
    base_name = f"{source_name_stem}_{type_val}"
    suffix_parts = []

    is_cog = params_dict.get(COL_COG)
    if is_cog:
        suffix_parts.append("cog" if is_cog else "tif")

    has_overviews = params_dict.get(COL_OVERVIEWS)
    if has_overviews:
        suffix_parts.append("ovr" if has_overviews else "flat")

    algo = params_dict.get(COL_ALGORITHM)
    if algo and str(algo).strip():
        suffix_parts.append(f"a{str(algo).upper().replace(ALGO_NONE, 'None')}") # Keep 'None' for filename consistency if that's existing behavior

    level = params_dict.get(COL_LEVEL)
    if level and str(level).strip() and algo in [ALGO_DEFLATE, ALGO_ZSTD]:
        suffix_parts.append(f"l{level}")

    predictor = params_dict.get(COL_PREDICTOR)
    if predictor and str(predictor).strip() and algo in [ALGO_LZW, ALGO_DEFLATE, ALGO_ZSTD]:
        suffix_parts.append(f"p{predictor}")

    decimals = params_dict.get(COL_DECIMALS)
    if decimals and str(decimals).strip() and type_val in [TYPE_DEM, TYPE_ERROR, TYPE_SCIENTIFIC]:
        suffix_parts.append(f"d{decimals}")

    max_z = params_dict.get(COL_MAX_Z_ERROR)
    if max_z and str(max_z).strip() and algo == ALGO_LERC:
        suffix_parts.append(f"m{max_z}")

    quality = params_dict.get(COL_QUALITY)
    if quality and str(quality).strip() and algo in [ALGO_JPEG, ALGO_JXL]:
        suffix_parts.append(f"q{quality}")

    if suffix_parts:
        base_name += "_" + "_".join(suffix_parts)

    return temp_dir_path / f"{base_name}.tif"

def _call_optimize_geotiff(
    source_tif_path: Path,
    output_tif_path: Path,
    type_val: str,
    optimize_params: Dict[str, Any],
    arc_mode: bool = False
) -> Tuple[float, bool, str]:
    """Calls the optimization script directly as a function."""
    
    # Convert optimize_params from the CSV into an OptimizeArguments object
    args = OptimizeArguments(
        input_path = source_tif_path,
        output_path = output_tif_path,
        product_type = type_val,
        algorithm = optimize_params.get(COL_ALGORITHM),
        vertical_srs = optimize_params.get(COL_VERTICAL_SRS, None),
        nodata = optimize_params.get(COL_NODATA, None),
        level = int(optimize_params[COL_LEVEL]) if optimize_params.get(COL_LEVEL) else None,
        predictor = int(optimize_params[COL_PREDICTOR]) if optimize_params.get(COL_PREDICTOR) else None,
        decimals = int(optimize_params[COL_DECIMALS]) if optimize_params.get(COL_DECIMALS) else None,
        max_z_error = float(optimize_params[COL_MAX_Z_ERROR]) if optimize_params.get(COL_MAX_Z_ERROR) else None ,
        quality = int(optimize_params[COL_QUALITY]) if optimize_params.get(COL_QUALITY) else None,
        tile_size = int(optimize_params.get(COL_TILE_SIZE, 512)),
        cog = str2bool(optimize_params.get(COL_COG, True)),
        overviews = str2bool(optimize_params.get(COL_OVERVIEWS, True)),
        open_report = False,
        geo_metadata=False,
        write_pam_xml = False,
        arc_mode = arc_mode
    )

    logger.info(f"  Calling optimization function with args: {args}")
    
    start_time = time.perf_counter()
    try:
        if arc_mode:
            logger.info("  Running in ArcGIS mode. Delegating to optimize_compression_arc.optimize_compression (subprocess wrapper).")
            # Use the arc-specific wrapper that handles subprocess execution in OSGeo4W
            return_code = optimize_compression_arc.optimize_compression(args)
        else:
             # Run normally in current process
            return_code = optimize_compression(args)

        end_time = time.perf_counter()
        duration = end_time - start_time
        
        if return_code != 0:
            error_message = f"Optimization function returned non-zero exit code: {return_code}"
            logger.error(error_message)
            return duration, False, error_message

        if not output_tif_path.exists():
            error_message = f"Output file was not created: {output_tif_path}"
            logger.error(error_message)
            return duration, False, error_message
            
        logger.debug(f"Optimization completed in {duration:.3f} seconds")
        return duration, True, ""

    except Exception as e:
        end_time = time.perf_counter()
        duration = end_time - start_time
        # Check if the exception is a ProcessingStepFailedError from our scripts
        # This is a bit tricky due to how modules are imported, so we check the class name.
        if type(e).__name__ == 'ProcessingStepFailedError':
            error_msg = str(e) # Pass the clean error message directly
            logger.error(f"A controlled processing step failed: {error_msg}")
        else:
            error_msg = f"Optimization function execution failed: {e}"
            logger.error(f"EXCEPTION in call_optimize_geotiff: {type(e).__name__}: {str(e)}\nTraceback:\n{''.join(traceback.format_tb(e.__traceback__))}")
        return duration, False, error_msg

def _call_optimize_geotiff_multiple(
    source_tif_path: Path,
    output_tif_path: Path,
    type_val: str,
    optimize_params: Dict[str, Any],
    arc_mode: bool = False,
    num_runs: Optional[int] = None
) -> Tuple[float, bool, str]:
    """Calls optimization script multiple times and averages the duration."""
    # Reduce runs for ArcGIS mode due to subprocess overhead
    if num_runs is None:
        num_runs = 1 if arc_mode else NUM_COMPRESSION_RUNS
    
    logger.info(f"  Running compression {num_runs} times and averaging...")
    
    durations = []
    all_success = True
    error_messages = []
    
    for run_num in range(num_runs):
        # Remove output file if it exists from previous run
        if Path(output_tif_path).exists():
            try:
                Path(output_tif_path).unlink()
            except OSError as e:
                error_messages.append(f"Could not delete existing output file on run {run_num+1}: {e}")
                all_success = False
                break
        
        logger.info(f"\n    Compression run {run_num+1}/{num_runs}...")
        duration, success, error_msg = _call_optimize_geotiff(
            source_tif_path, output_tif_path, type_val, optimize_params, arc_mode
        )
        
        if success:
            durations.append(duration)
        else:
            error_messages.append(f"Run {run_num+1}: {error_msg}")
            all_success = False
            # It's important to break here so we don't try to average durations if a run failed.
            break # Ensure we break out of the loop on first failure.
    
    # After the loop, check conditions for returning success or failure
    if all_success and durations:
        avg_duration = sum(durations) / len(durations)
        if len(durations) > 1:
            # Sample standard deviation
            std_dev = (sum((d - avg_duration) ** 2 for d in durations) / (len(durations) - 1)) ** 0.5
            logger.info(f"    Average compression time: {avg_duration:.3f}s (±{std_dev:.3f}s)")
        else: # Only one successful run
            logger.info(f"    Compression time (1 run): {avg_duration:.3f}s")
        return avg_duration, True, ""
    elif not durations and all_success: # All runs might have been skipped or had an issue before duration append
        return 0, False, "No successful compression runs to average, though no explicit errors reported during runs."
    else: # Some error occurred, or no durations were recorded
        combined_error = "; ".join(error_messages)
        if not combined_error and not durations: # Fallback error message
            combined_error = "Compression failed or no durations recorded for unknown reasons."
        return 0, False, combined_error

def _measure_read_speed(
    file_path: Path,
    file_size_mb: float,
    num_warmup_iterations: int = NUM_WARMUP_READ_ITERATIONS,
    num_measured_iterations: int = NUM_MEASURED_READ_ITERATIONS
) -> Tuple[float, str]:
    """Measures GeoTIFF read speed after warm-up, returning average MB/s."""
    if not Path(file_path).exists():
        return 0.0, f"File not found for reading: {file_path}"
    if file_size_mb <= 0: # Check after ensuring file_path is valid Path object
        return 0.0, "File size is zero or negative, cannot calculate read speed."

    # Warm-up iterations to get data into OS cache
    logger.info(f"    Performing {num_warmup_iterations} warm-up read iterations for {file_path.name}...")
    for i in range(num_warmup_iterations):
        ds = None
        try:
            ds = gdal.Open(str(file_path), gdal.GA_ReadOnly)
            if ds is None:
                err_msg = f"GDAL.Open failed for {file_path} during warm-up iteration {i+1}."
                logger.warning(err_msg)
                return 0.0, err_msg
            
            for band_idx in range(1, ds.RasterCount + 1):
                band = ds.GetRasterBand(band_idx)
                if band is None:
                    err_msg = f"GetRasterBand failed for band {band_idx} in {file_path} during warm-up iteration {i+1}."
                    logger.warning(err_msg)
                    return 0.0, err_msg
                buf_xsize = min(band.XSize, 1024)
                buf_ysize = min(band.YSize, 1024)
                _ = band.ReadRaster(0, 0, band.XSize, band.YSize, buf_xsize=buf_xsize, buf_ysize=buf_ysize)
        except Exception as e:
            err_msg = f"Exception during GDAL warm-up read for {file_path} on iteration {i+1}: {e}"
            logger.warning(err_msg)
            return 0.0, err_msg
        finally:
            band = None
            ds = None

    # Now perform measured iterations
    logger.info(f"    Measuring {num_measured_iterations} read iterations for {file_path.name}...")
    total_read_time_seconds = 0.0
    error_occurred_msg = None

    for i in range(num_measured_iterations):
        ds = None
        try:
            iter_start_time = time.perf_counter()
            ds = gdal.Open(str(file_path), gdal.GA_ReadOnly)
            if ds is None:
                error_occurred_msg = f"GDAL.Open failed for {file_path} on measured iteration {i+1}."
                logger.warning(error_occurred_msg)
                break

            for band_idx in range(1, ds.RasterCount + 1):
                band = ds.GetRasterBand(band_idx)
                if band is None:
                    error_occurred_msg = f"GetRasterBand failed for band {band_idx} in {file_path} on measured iteration {i+1}."
                    logger.warning(error_occurred_msg)
                    break
                buf_xsize = min(band.XSize, 1024)
                buf_ysize = min(band.YSize, 1024)
                _ = band.ReadRaster(0, 0, band.XSize, band.YSize, buf_xsize=buf_xsize, buf_ysize=buf_ysize)
            
            if error_occurred_msg:
                break

            iter_end_time = time.perf_counter()
            total_read_time_seconds += (iter_end_time - iter_start_time)

        except Exception as e:
            error_occurred_msg = f"Exception during GDAL read for {file_path} on measured iteration {i+1}: {e}"
            logger.warning(error_occurred_msg)
            break
        finally:
            band = None
            ds = None

    if error_occurred_msg:
        # Already logged inside the loop if it's a specific GDAL error.
        # This catches the case where the loop finishes due to error_occurred_msg.
        return 0.0, error_occurred_msg

    if num_measured_iterations == 0:
        return 0.0, "Number of measured iterations for read speed was zero."
    if total_read_time_seconds <= 0: # Check for non-positive total time
         return 0.0, "Total read time was zero or negative. File might be too small or iterations too fast."

    avg_read_time_seconds = total_read_time_seconds / num_measured_iterations
    # It's highly unlikely avg_read_time_seconds would be zero if total_read_time_seconds > 0
    # and num_measured_iterations > 0, but good to be safe.
    if avg_read_time_seconds <= 0:
        return 0.0, "Average read time was zero or negative."
        
    read_speed_mb_s = file_size_mb / avg_read_time_seconds
    return read_speed_mb_s, ""


@dataclass
class GeoTiffMetadata:
    algorithm: str = ALGO_NONE
    tile_size: Optional[int] = None
    predictor: Optional[int] = None
    level: Optional[int] = None
    quality: Optional[int] = None
    decimals: Optional[int] = None
    max_z_error: Optional[float] = None
    size_mb: float = 0.0
    gdal_type_name: str = VAL_NA
    error: Optional[str] = None


def _get_geotiff_metadata(filepath: Path) -> GeoTiffMetadata:
    """Opens a GeoTIFF and extracts relevant metadata."""
    metadata = GeoTiffMetadata()
    ds = None
    try:
        if not Path(filepath).exists():
            metadata.error = f"File not found: {filepath}"
            return metadata
            
        ds = gdal.Open(str(filepath), gdal.GA_ReadOnly)
        if ds is None:
            metadata.error = f"GDAL.Open failed for {filepath}"
            return metadata

        metadata.tile_size = _get_tile_size(ds)
        
        metadata.size_mb = Path(filepath).stat().st_size / (1024 * 1024)

        algo_meta = ds.GetMetadataItem('COMPRESSION', 'IMAGE_STRUCTURE')
        metadata.algorithm = str(algo_meta).upper() if algo_meta else ALGO_NONE

        pred_meta = ds.GetMetadataItem('PREDICTOR', 'IMAGE_STRUCTURE')
        if pred_meta:
            metadata.predictor = int(pred_meta)
        
        # Attempt to get level for common types if possible (example, may need refinement)
        # This part is tricky as GDAL doesn't expose all TIFF tags easily.
        # For DEFLATE/ZSTD, it might be in TIFF metadata.
        tiff_metadata = ds.GetMetadata("TIFF")
        if metadata.algorithm == ALGO_DEFLATE and tiff_metadata and 'TIFFTAG_ZLEVEL' in tiff_metadata:
            metadata.level = int(tiff_metadata['TIFFTAG_ZLEVEL'])
        elif metadata.algorithm == ALGO_ZSTD and tiff_metadata and 'ZSTD_LEVEL' in tiff_metadata: # Assuming 'ZSTD_LEVEL' tag
             metadata.level = int(tiff_metadata['ZSTD_LEVEL'])

        # Quality handling for JPEG and JXL
        if metadata.algorithm in [ALGO_JPEG, ALGO_JXL]:
            # Quality is not stored in the TIFF header for JPEG/JXL, so it will be VAL_NA for the original.
            # For compressed files, it should be populated from the input CSV.
            # The format_output_row function will handle displaying the correct value.
            metadata.quality = None
        
        if metadata.algorithm == ALGO_LERC:
            metadata.max_z_error = 0.0
            max_z_meta = ds.GetMetadataItem('MAX_Z_ERROR', 'IMAGE_STRUCTURE')
            if max_z_meta:
                metadata.max_z_error = float(max_z_meta)

        if ds.RasterCount > 0:
            band = ds.GetRasterBand(1)
            dt = band.DataType
            metadata.gdal_type_name = gdal.GetDataTypeName(dt)
            if dt in [gdal.GDT_Float32, gdal.GDT_Float64]:
                # Attempt to detect decimal precision for float types
                precision_result = determine_decimal_precision(ds)
                if isinstance(precision_result, list):
                    metadata.decimals = precision_result[0] if precision_result else None
                else:
                    metadata.decimals = precision_result
        
    except Exception as e:
        metadata.error = f"Exception in get_geotiff_metadata for {filepath}: {e}"
    finally:
        band = None
        ds = None
    return metadata


def _calculate_speed(file_size_mb: float, time_seconds: float) -> float:
    """Calculates speed in MB/s. Returns 0 if time_seconds is 0 or None."""
    if not time_seconds or time_seconds <= 0: # Ensure time is positive
        return 0.0
    return file_size_mb / time_seconds

# --- Main Processing Logic ---

def _read_test_cases_from_csv(csv_path: Path) -> List[Dict[str, str]]:
    """Reads compression test cases from the input CSV file."""
    test_cases: List[Dict[str, str]] = []
    try:
        with open(csv_path, 'r', newline='', encoding='utf-8') as infile:
            reader = csv.DictReader(infile)
            if not reader.fieldnames:
                raise CSVLoadError(f"Input CSV {csv_path} has no headers.")
            if not all(header in reader.fieldnames for header in csv_file_MANDATORY_HEADERS):
                missing = [h for h in csv_file_MANDATORY_HEADERS if h not in reader.fieldnames]
                raise CSVLoadError(f"Input CSV {csv_path} is missing essential headers: {', '.join(missing)}")
            for row in reader:
                test_cases.append(row)
        if not test_cases:
            raise CSVLoadError(f"Input CSV file {csv_path} is empty or contains no data rows.")
    except FileNotFoundError:
        raise CSVLoadError(f"Input CSV file not found: {csv_path}")
    except CSVLoadError:
        raise  # Re-raise CSVLoadError
    except Exception as e:
        raise CSVLoadError(f"Error reading input CSV {csv_path}: {e}")
    return test_cases


class ExcelWriter:
    """Handles writing results to an Excel workbook."""
    def __init__(self, excel_path: Path, start_row: int = 3):
        self.excel_path = excel_path
        self.start_row = start_row
        self.current_row = start_row
        self.group_start_row = None  # Track the start of each file group
        try:
            template_path = resources.files('gttk.resources.templates').joinpath('test_compression_template.xlsx')
            with resources.as_file(template_path) as template_file:
                self.workbook = load_workbook(template_file)

            self.worksheet = self.workbook.active

        except Exception as e:
            logger.error(f"Could not load Excel template {excel_path}: {e}")
            sys.exit(1)

    def _get_merge_anchor(self, cell):
        """ Returns the top-left anchor cell if this cell is within a merged range."""
        ws = self.worksheet
        if not ws:
            return cell
        
        for r in ws.merged_cells.ranges:
            if cell.coordinate in r:
                return ws.cell(row=r.min_row, column=r.min_col)
        return cell

    def _apply_template_style_to_row(self, dst_row: int, template_row: Optional[int] = None):
        """Preserves the Excel template styles in the destination row."""
        if not self.worksheet:
            return
        
        ws = self.worksheet
        src_row = template_row or self.start_row  # usually 3

        for col_idx in range(1, len(OUTPUT_COLUMNS) + 1):
            src = ws.cell(row=src_row, column=col_idx)
            dst = ws.cell(row=dst_row, column=col_idx)

            # Resolve anchors so we don't try to style a MergedCell
            src_eff = self._get_merge_anchor(src)
            dst_eff = self._get_merge_anchor(dst)

            # Pylance types these as StyleProxy on get; cast them to concrete types for assignment
            dst_eff.font = cast(Font, copy(src_eff.font))
            dst_eff.fill = cast(PatternFill, copy(src_eff.fill))
            dst_eff.border = cast(Border, copy(src_eff.border))
            dst_eff.alignment = cast(Alignment, copy(src_eff.alignment))
            dst_eff.protection = cast(Protection, copy(src_eff.protection))
            dst_eff.number_format = src_eff.number_format

    def mark_group_start(self):
        """Mark the start of a new file group."""
        self.group_start_row = self.current_row

    def mark_group_end(self):
        """Apply a thick bottom border to the last row in the group, preserving existing left/right/top."""
        if self.group_start_row is None or self.worksheet is None:
            return

        # Only apply thick bottom border to the last row of the group (current_row - 1)
        last_row = self.current_row - 1
        thick_bottom = Side(border_style="thick", color="000000")  # Side

        for col_idx in range(1, len(OUTPUT_COLUMNS) + 1):
            cell = self.worksheet.cell(row=last_row, column=col_idx)
            existing = cell.border if cell.border else Border()
            # Preserve left, right, and top; only update bottom to thick
            cell.border = Border(
                left=existing.left,
                right=existing.right,
                top=existing.top,
                bottom=thick_bottom
            )

        # Reset group start for next group
        self.group_start_row = None

    def write_row(self, row_data: Dict[str, Any]):
        """Writes a dictionary of data to the current row in the worksheet."""
        if not self.worksheet:
            return
        
        # Prime the row with template styles
        self._apply_template_style_to_row(self.current_row, template_row=self.start_row)

        for col_idx, column_name in enumerate(OUTPUT_COLUMNS, start=1):
            raw_value = row_data.get(column_name, '')

            # Make boolean 'COG' or 'Overviews' values easier to distinguish
            if column_name in ['cog', 'overviews']:
                if raw_value == '':
                    value = ''
                else:
                    value = '✅ Yes' if bool(raw_value) else '❎ No'
            else:
                # Coerce simple numeric strings to numbers; keep as-is on failure
                if isinstance(raw_value, str):
                    try:
                        v = raw_value.strip()
                        if v:
                            if v.lstrip('-').isdigit():
                                raw_value = int(v)  # handles negative integers
                            else:
                                raw_value = float(v)
                    except ValueError:
                        pass
                value = raw_value

            cell = self.worksheet.cell(row=self.current_row, column=col_idx)
            cell_eff = self._get_merge_anchor(cell)
            cast(Any, cell_eff).value = value

        self.current_row += 1

    def save(self, success: bool = True):
        """Saves the workbook with controlled print area, selection, and page setup."""
        if not self.worksheet:
            logger.error("Cannot save, no active worksheet.")
            return

        try:
            ws = self.worksheet
            
            # --- Finalize Worksheet Formatting ---
            # 1. Define Print Area to avoid printing thousands of blank pages
            last_data_row = max(self.start_row, self.current_row - 1)
            last_col_letter = get_column_letter(len(OUTPUT_COLUMNS))
            ws.print_area = f'A1:{last_col_letter}{last_data_row}'
            
            # 2. Repeat header rows 1 and 2 on each printed page
            ws.print_title_rows = '1:2'
            
            # 3. Set page setup to fit to 1 page wide, unlimited height
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0 # 0 means unlimited pages vertically
            
            # 4. Set the active cell to the first data cell (e.g., C3)
            start_cell = f'B{self.start_row}'
            ws.sheet_view.tabSelected = True
            if hasattr(ws.sheet_view, 'selection'):
                for sel in ws.sheet_view.selection:
                    sel.activeCell = start_cell
                    sel.sqref = start_cell

            # --- Save the Workbook ---
            self.workbook.save(self.excel_path)
            
            if success:
                logger.info(f"Results successfully saved to: {self.excel_path.resolve()}")
            else:
                logger.info(f"Partial results saved to: {self.excel_path.resolve()}")

        except Exception as e:
            logger.error(f"Could not save Excel file {self.excel_path}: {e}")


@dataclass
class TestResultMetrics:
    """Holds metrics for a single compression test or baseline."""
    tile_size: Optional[int] = None
    size_mb: Optional[float] = None
    compression_pct: Optional[float] = None
    write_speed_mb_s: Optional[float] = None
    read_speed_mb_s: Optional[float] = None
    filepath: Optional[Path] = None
    processing_error_msg: str = VAL_NA
    success_flag: bool = False


def _process_geotiff_for_metrics(
    args: TestArguments,
    source_path: Path,
    output_path: Path,
    processing_params: Dict[str, Any],
    is_baseline: bool = False
) -> TestResultMetrics:
    """Processes a GeoTIFF (original or generated) to gather performance metrics."""
    metrics = TestResultMetrics(filepath=output_path)
    file_type = processing_params.get(COL_TYPE, VAL_NA)

    if is_baseline: # For the script-generated 'NONE' baseline
        logger.info(f"  Generating reference '{ALGO_NONE}' file: {output_path}")
        if not args.optimize_script_path:
            raise OptimizationError("Optimization script path is not set")
        assert args.optimize_script_path is not None, "optimize_script_path must be set"
        write_duration_s, success, call_error_msg = _call_optimize_geotiff_multiple(
            source_path, output_path, file_type, processing_params, args.arc_mode or False
        )
    else: # For regular test cases
        logger.info(f"  Generating compressed file: {output_path}")
        if not args.optimize_script_path:
            raise OptimizationError("Optimization script path is not set")
        assert args.optimize_script_path is not None, "optimize_script_path must be set"
        write_duration_s, success, call_error_msg = _call_optimize_geotiff_multiple(
            source_path, output_path, file_type, processing_params, args.arc_mode or False
        )
        if not success:
            metrics.processing_error_msg += f"OptimizeCallError: {call_error_msg}; "
    
    if not output_path.exists():
        metrics.processing_error_msg += f"FileNotGenerated: {output_path.name} was not found post-processing attempt; "
        success = False # Ensure success is false if file doesn't exist
    
    if success:
        try:
            metrics.tile_size = processing_params.get(COL_TILE_SIZE, VAL_NA)
            if not metrics.tile_size or metrics.tile_size == VAL_NA:
                # If tile size was not in params (e.g. baseline), try to read it from the generated file
                try:
                    ds_temp = gdal.Open(str(output_path), gdal.GA_ReadOnly)
                    if ds_temp:
                         ts = _get_tile_size(ds_temp)
                         if ts is not None:
                             metrics.tile_size = int(ts)
                    ds_temp = None
                except Exception:
                    pass

            metrics.size_mb = output_path.stat().st_size / (1024 * 1024)
            metrics.compression_pct = calculate_compression_efficiency(str(output_path), debug=True)
            if is_baseline or write_duration_s > 0 : # write_duration_s is from optimize_geotiff call
                 metrics.write_speed_mb_s = _calculate_speed(metrics.size_mb, write_duration_s)
            
            read_speed, read_err = _measure_read_speed(output_path, metrics.size_mb)
            if read_err:
                metrics.processing_error_msg += f"ReadError: {read_err}; "
                metrics.read_speed_mb_s = 0.0 # Indicate error
            else:
                metrics.read_speed_mb_s = read_speed
        except FileNotFoundError as e_fnf:
            metrics.processing_error_msg += f"FileNotFoundDuringStats: {e_fnf}; "
            success = False
        except Exception as e_stats:
            metrics.processing_error_msg += f"StatsError: {e_stats}; "
            success = False
    else: # If optimize_geotiff_multiple failed (success=False)
        if not metrics.processing_error_msg: # Ensure there's an error message if not already set
            metrics.processing_error_msg = call_error_msg or "Unknown optimization error."

    if metrics.size_mb is None or metrics.size_mb <= 0:
        metrics.processing_error_msg += f"InvalidFileSize: Size is {metrics.size_mb}MB; "
        success = False # Mark as overall failure for this processing step

    # Return success based on whether critical metrics could be gathered
    metrics.success_flag = success and metrics.size_mb is not None and metrics.size_mb > 0
    return metrics

def _format_output_row(
    filename: str,
    params: Dict[str, Any], # Could be from test_case or derived for original/baseline
    current_metrics: TestResultMetrics,
    baseline_metrics: TestResultMetrics,
    original_size_mb: Optional[float],
    original_compression_pct: Optional[float]
) -> Dict[str, Any]:
    """Formats a single row for Excel output."""
    row: Dict[str, Any] = {k: VAL_NA for k in OUTPUT_COLUMNS}  # Initialize with N/A
    
    # Populate from params
    row[COL_FILENAME] = filename
    for key in [COL_TYPE, COL_TILE_SIZE, COL_ALGORITHM, COL_LEVEL, COL_PREDICTOR, COL_DECIMALS, COL_MAX_Z_ERROR, COL_QUALITY]:
        if key in params:
            row[key] = params[key]

    # Validate COG and overviews
    if current_metrics.filepath and current_metrics.filepath.exists():
        row[COL_COG] = _is_cog(current_metrics.filepath)
        row[COL_OVERVIEWS] = _has_internal_overviews(current_metrics.filepath)
    else:
        row[COL_COG] = ''
        row[COL_OVERVIEWS] = ''

    # If tile size is not in params (e.g. baseline or detected from file), try to use metrics
    if not row.get(COL_TILE_SIZE) and current_metrics.tile_size is not None:
        row[COL_TILE_SIZE] = current_metrics.tile_size

    error_accumulator = current_metrics.processing_error_msg.strip()

    if current_metrics.success_flag and current_metrics.size_mb is not None:
        row[COL_SIZE_MB] = current_metrics.size_mb
        
        # Calculate size delta percentage vs original file
        if original_size_mb and original_size_mb > 0:
            size_delta_pct = ((current_metrics.size_mb - original_size_mb) / original_size_mb) * 100
            row[COL_SIZE_DELTA] = size_delta_pct
        else:
            row[COL_SIZE_DELTA] = VAL_ERROR

        row[COL_COMPRESSION_PERCENT] = current_metrics.compression_pct

        # Improvement vs Original Source: difference between current and original compression
        if params.get(COL_ALGORITHM) == ALGO_NONE:
            row[COL_SIZE_DELTA] = ''  # Baseline compression is irrelevant
            row[COL_IMPROVEMENT_PERCENT] = ''  # Baseline has no improvement to measure
        elif original_compression_pct is not None and current_metrics.compression_pct is not None:
            improvement_pct = current_metrics.compression_pct - original_compression_pct
            row[COL_IMPROVEMENT_PERCENT] = improvement_pct
        else:
            row[COL_IMPROVEMENT_PERCENT] = VAL_ERROR
            if original_compression_pct is None:
                error_accumulator += "Original compression % unavailable for improvement calculation; "
        
        # Write Speed
        if current_metrics.write_speed_mb_s is not None:
            row[COL_WRITE_SPEED] = current_metrics.write_speed_mb_s
            if baseline_metrics.write_speed_mb_s and baseline_metrics.write_speed_mb_s > 0:
                write_ratio = current_metrics.write_speed_mb_s / baseline_metrics.write_speed_mb_s
                row[COL_WRITE_RATIO] = write_ratio
            else:
                row[COL_WRITE_RATIO] = VAL_ERROR
                error_accumulator += "Baseline write speed invalid for ratio; "
        else:  # No write speed if it's the original file, or if processing failed before write speed calc
            row[COL_WRITE_SPEED] = VAL_NA if params.get(COL_TYPE) == TYPE_ORIGINAL else VAL_ERROR
            row[COL_WRITE_RATIO] = VAL_NA if params.get(COL_TYPE) == TYPE_ORIGINAL else VAL_ERROR

        # Read Speed
        if current_metrics.read_speed_mb_s is not None:
            row[COL_READ_SPEED] = current_metrics.read_speed_mb_s
            if baseline_metrics.read_speed_mb_s and baseline_metrics.read_speed_mb_s > 0:
                read_ratio = current_metrics.read_speed_mb_s / baseline_metrics.read_speed_mb_s
                row[COL_READ_RATIO] = read_ratio
            else:
                row[COL_READ_RATIO] = VAL_ERROR
                error_accumulator += "Baseline read speed invalid for ratio; "
        else:
            row[COL_READ_SPEED] = VAL_ERROR
            row[COL_READ_RATIO] = VAL_ERROR
            error_accumulator += "Read speed data missing; "

    else:  # current_metrics.success_flag is False
        for k_metric in [COL_SIZE_MB, COL_SIZE_DELTA, COL_COMPRESSION_PERCENT, COL_IMPROVEMENT_PERCENT,
                         COL_WRITE_SPEED, COL_WRITE_RATIO, COL_READ_SPEED, COL_READ_RATIO]:
            if not row.get(k_metric) or row.get(k_metric) == VAL_NA:  # Don't overwrite if already set by params
                row[k_metric] = VAL_ERROR if not error_accumulator else VAL_SKIPPED  # Skipped if error implies no processing
        if not error_accumulator:
            error_accumulator = "Processing failed, metrics unavailable."
    
    row[COL_COMMENT] = error_accumulator.strip()
    return row


def _process_original_source_file(source_file: Path, args: TestArguments, excel_writer: ExcelWriter) -> Tuple[Optional[float], Optional[float]]:
    """Processes the original source GeoTIFF and writes its row to Excel.
    
    Returns:
        The original file's compression percentage (for improvement calculations), or None if processing failed.
    """
    logger.info(f"\nProcessing Original Source File: {source_file.name}...")
    
    source_file_metadata = _get_geotiff_metadata(source_file)
    original_metrics = TestResultMetrics(
        tile_size=source_file_metadata.tile_size,
        size_mb=source_file_metadata.size_mb,
        filepath=source_file,
        processing_error_msg=source_file_metadata.error or ""
    )

    # Measure read speed for original source
    if original_metrics.size_mb and original_metrics.size_mb > 0:
        read_speed, read_err = _measure_read_speed(source_file, original_metrics.size_mb)
        if read_err:
            original_metrics.processing_error_msg += f"ReadError: {read_err}; "
            original_metrics.read_speed_mb_s = 0.0
        else:
            original_metrics.read_speed_mb_s = read_speed
    else:
        original_metrics.processing_error_msg += "Original source size invalid for read speed measurement; "
        original_metrics.read_speed_mb_s = 0.0

    # Original file doesn't have a "write speed" in this context
    original_metrics.write_speed_mb_s = None
    original_metrics.success_flag = original_metrics.size_mb is not None and original_metrics.size_mb > 0

    # Calculate compression efficiency using TIFF tags (consistent with report_helpers logic)
    original_compression_pct = None
    if original_metrics.success_flag:
        original_compression_pct = calculate_compression_efficiency(str(source_file), debug=True)

    # Prepare parameters for formatting (mimicking a test_case structure)
    original_params = {
        COL_TYPE: TYPE_ORIGINAL,
        COL_ALGORITHM: source_file_metadata.algorithm,
        COL_TILE_SIZE: source_file_metadata.tile_size,
        COL_PREDICTOR: source_file_metadata.predictor,
        COL_LEVEL: source_file_metadata.level,
        COL_QUALITY: source_file_metadata.quality,
        COL_DECIMALS: source_file_metadata.decimals,
        COL_MAX_Z_ERROR: source_file_metadata.max_z_error,
    }

    # For original source row, don't compare against baseline - use special handling
    output_row_data: Dict[str, Any] = {k: VAL_NA for k in OUTPUT_COLUMNS}
    output_row_data[COL_FILENAME] = source_file.name
    for key in [COL_TYPE, COL_ALGORITHM, COL_TILE_SIZE, COL_PREDICTOR, COL_LEVEL, COL_QUALITY, COL_DECIMALS, COL_MAX_Z_ERROR]:
        if key in original_params:
            output_row_data[key] = original_params[key]

    # Validate COG and overviews for the original file
    output_row_data[COL_COG] = _is_cog(source_file)
    output_row_data[COL_OVERVIEWS] = _has_internal_overviews(source_file)
    
    if original_metrics.success_flag and original_metrics.size_mb is not None:
        output_row_data[COL_SIZE_MB] = original_metrics.size_mb
        output_row_data[COL_SIZE_DELTA] = 0.0  # Original file is the baseline
        
        output_row_data[COL_COMPRESSION_PERCENT] = original_compression_pct if original_compression_pct is not None else ""
        
        output_row_data[COL_IMPROVEMENT_PERCENT] = 0.0
        output_row_data[COL_WRITE_SPEED] = VAL_NA  # No write operation
        output_row_data[COL_WRITE_RATIO] = VAL_NA  # No write operation
        
        if original_metrics.read_speed_mb_s is not None:
            output_row_data[COL_READ_SPEED] = original_metrics.read_speed_mb_s
            output_row_data[COL_READ_RATIO] = VAL_NA  # Original is the reference
        else:
            output_row_data[COL_READ_SPEED] = VAL_ERROR
            output_row_data[COL_READ_RATIO] = VAL_ERROR
    else:
        for k_metric in [COL_SIZE_MB, COL_SIZE_DELTA, COL_COMPRESSION_PERCENT, COL_IMPROVEMENT_PERCENT,
                         COL_WRITE_SPEED, COL_WRITE_RATIO, COL_READ_SPEED, COL_READ_RATIO]:
            output_row_data[k_metric] = VAL_ERROR
    
    output_row_data[COL_COMMENT] = original_metrics.processing_error_msg.strip()
    
    excel_writer.write_row(output_row_data)
    return original_metrics.size_mb, original_compression_pct


def _process_script_baseline(source_file: Path, args: TestArguments, excel_writer: ExcelWriter, first_test_case_type: str, original_size_mb: Optional[float], original_compression_pct: Optional[float]) -> TestResultMetrics:
    """Processes the script-generated uncompressed baseline and writes its row, returning its metrics."""
    logger.info(f"\nProcessing Script-Generated Uncompressed Baseline ('{ALGO_NONE}' reference for ratios)...")
    baseline_processing_params = {
        COL_TYPE: 'baseline', # Use type from first test case for consistency
        COL_ALGORITHM: ALGO_NONE,
        'vertical_srs': 'EGM2008'  # Add vertical SRS for DEM processing; it doesn't matter which one for testing
    }
    assert args.temp_dir is not None, "temp_dir must be set"
    baseline_temp_filepath = _generate_temp_filename(
        source_file.stem,
        baseline_processing_params[COL_TYPE],
        {COL_ALGORITHM: ALGO_NONE}, # Params for filename generation
        args.temp_dir
    )

    current_baseline_metrics = _process_geotiff_for_metrics(
        args,
        source_path=source_file,
        output_path=baseline_temp_filepath,
        processing_params=baseline_processing_params,
        is_baseline=True
    )

    if not current_baseline_metrics.success_flag:
        critical_error_msg = f"CRITICAL: Reference uncompressed file processing failed or size invalid ({current_baseline_metrics.processing_error_msg.strip()}). Cannot proceed. Exiting."
        logger.error(critical_error_msg)
        error_row_data = {k: VAL_ERROR for k in OUTPUT_COLUMNS}
        error_row_data[COL_FILENAME] = baseline_temp_filepath.name if baseline_temp_filepath else "baseline_gen_failed"
        error_row_data[COL_COMMENT] = critical_error_msg
        excel_writer.write_row(error_row_data)
        excel_writer.save(success=False) # Attempt to save before exiting
        sys.exit(1)
    
    logger.info(f"  Baseline Metrics: Size={current_baseline_metrics.size_mb:.3f}MB, WriteSpeed={current_baseline_metrics.write_speed_mb_s:.2f}MB/s, ReadSpeed={current_baseline_metrics.read_speed_mb_s:.2f}MB/s")

    # Format baseline row for Excel
    # The baseline_metrics for format_output_row will be itself, leading to ratios of 1.00
    baseline_output_row = _format_output_row(
        filename=baseline_temp_filepath.name,
        params=baseline_processing_params,
        current_metrics=current_baseline_metrics,
        baseline_metrics=current_baseline_metrics,
        original_size_mb=original_size_mb,
        original_compression_pct=original_compression_pct
    )
    # Specific overrides for baseline row
    baseline_output_row[COL_COMPRESSION_PERCENT] = "0.0"  # By definition
    baseline_output_row[COL_WRITE_RATIO] = "1.00"
    baseline_output_row[COL_READ_RATIO] = "1.00"
    # Delta Size % and Improvement % for baseline are left blank (handled in _format_output_row)

    excel_writer.write_row(baseline_output_row)

    if args.delete_test_files and baseline_temp_filepath.exists():
        try:
            logger.info(f"  Deleting baseline temporary file: {baseline_temp_filepath}")
            baseline_temp_filepath.unlink()
        except OSError as e_del_ref:
            logger.warning(f"  Could not delete baseline file {baseline_temp_filepath}: {e_del_ref}")
            
    return current_baseline_metrics


def _run_compression_tests(source_file: Path, args: TestArguments, test_cases: List[Dict[str, str]], excel_writer: ExcelWriter, baseline_metrics: TestResultMetrics, original_size_mb: Optional[float], original_compression_pct: Optional[float]):
    """Runs all compression test cases from the CSV and writes results to Excel."""
    logger.info(f"\nProcessing {len(test_cases)} test cases for {source_file.name} from {args.csv_path} (comparing to generated '{ALGO_NONE}' reference)...")
    
    # Get the source file's data type once to use for validation in the loop
    source_metadata = _get_geotiff_metadata(source_file)
    is_source_float = 'Float' in source_metadata.gdal_type_name

    for i, test_params_csv_row in enumerate(test_cases):
        logger.info(f"\n--- Test Case {i+1}/{len(test_cases)} for {source_file.name} ---")
        logger.info(f"  Parameters: {test_params_csv_row}")

        current_type = test_params_csv_row.get(COL_TYPE)
        current_algo = test_params_csv_row.get(COL_ALGORITHM)
        row_error_accumulator = VAL_NA
        
        # Initialize output_row with defaults from test_params_csv_row
        output_row_data = {col: test_params_csv_row.get(col, '') for col in OUTPUT_COLUMNS}

        # Correct for NaN NoData value with non-float data types.
        if not is_source_float and test_params_csv_row.get(COL_NODATA) and str(test_params_csv_row[COL_NODATA]).lower() == 'nan':
            test_params_csv_row[COL_NODATA] = '' # Silently unset by using an empty string

        # Correct for Predictor=3 with integer data.
        if not is_source_float and test_params_csv_row.get(COL_PREDICTOR) == '3':
            error_msg = f"Predictor 3 is not valid for integer data type ('{source_metadata.gdal_type_name}')."
            error_row_data = _format_output_row(
                filename='',  # not created
                params=test_params_csv_row,
                current_metrics=TestResultMetrics(processing_error_msg=error_msg),
                baseline_metrics=baseline_metrics,
                original_size_mb=original_size_mb,
                original_compression_pct=original_compression_pct
            )
            excel_writer.write_row(error_row_data)
            logger.warning(f"Skipping test case: {error_msg}")
            continue

        if not current_type or not str(current_type).strip():
            row_error_accumulator = f"Missing '{COL_TYPE}' in input CSV row."
        elif not current_algo or not str(current_algo).strip():
            row_error_accumulator = f"Missing '{COL_ALGORITHM}' in input CSV row."

        if row_error_accumulator:
            logger.warning(f"  Skipping test case {i+1}: {row_error_accumulator}")
            output_row_data[COL_COMMENT] = row_error_accumulator
            for k_metric in [COL_SIZE_MB, COL_SIZE_DELTA, COL_COMPRESSION_PERCENT, COL_IMPROVEMENT_PERCENT,
                             COL_WRITE_SPEED, COL_WRITE_RATIO, COL_READ_SPEED, COL_READ_RATIO]:
                output_row_data[k_metric] = VAL_SKIPPED
            excel_writer.write_row(output_row_data)
            continue

        # Ensure current_type is a string
        type_for_filename = str(current_type) if current_type else "unknown"
        
        assert args.temp_dir is not None, "temp_dir must be set"
        temp_filepath = _generate_temp_filename(
            source_file.stem,
            type_for_filename,
            test_params_csv_row, # Pass full dict for filename generation
            args.temp_dir
        )
        output_row_data[COL_FILENAME] = temp_filepath.name # Set filename early

        current_test_metrics = _process_geotiff_for_metrics(
            args,
            source_path=source_file,
            output_path=temp_filepath,
            processing_params=test_params_csv_row, # Pass full dict for processing
            is_baseline=False
        )
        
        # Format the rest of the row using the gathered metrics
        formatted_row = _format_output_row(
            filename=temp_filepath.name,
            params=test_params_csv_row,
            current_metrics=current_test_metrics,
            baseline_metrics=baseline_metrics,
            original_size_mb=original_size_mb,
            original_compression_pct=original_compression_pct
        )
        excel_writer.write_row(formatted_row)

        if temp_filepath.exists():
            if current_test_metrics.success_flag and args.delete_test_files:
                try:
                    logger.info(f"  Deleting temporary file: {temp_filepath}")
                    gc.collect() # Force garbage collection to release any lingering file handles
                    temp_filepath.unlink()
                    logger.info(f"  Successfully deleted: {temp_filepath}")
                except OSError as e_del_loop:
                    logger.warning(f"  Could not delete temp file {temp_filepath}: {e_del_loop}")
            elif not current_test_metrics.success_flag:
                action = "not deleted due to error (delete_test_files=True)" if args.delete_test_files else "not deleted (delete_test_files=False)"
                logger.info(f"  Note: Test for {temp_filepath.name} failed. File {action}. Path: {temp_filepath}")


def test_compression(args: TestArguments):
    """Main function to orchestrate the GeoTIFF compression testing."""
    # --- Logging Setup ---
    assert args.temp_dir is not None, "temp_dir must be set"
    log_file_path = args.log_file or args.temp_dir / "test_compression_debug.log"
    log_file_path.parent.mkdir(parents=True, exist_ok=True)
    
    # Configure FileHandler for logging
    file_handler = logging.FileHandler(log_file_path, mode='w', encoding='utf-8')
    file_handler.setFormatter(logging.Formatter('%(asctime)s | %(levelname)s | %(funcName)s | %(lineno)d | %(message)s'))
    logger.addHandler(file_handler)
    logger.setLevel(logging.DEBUG)

    arcMode = args.arc_mode or False
    if arcMode:
        init_arcpy()
        # Add Arcpy handler for real-time feedback in ArcGIS Pro
        try:
            arcpy_handler = ArcpyLogHandler()
            arcpy_handler.setFormatter(logging.Formatter('%(message)s'))
            logger.addHandler(arcpy_handler)
        except Exception as e:
            logger.warning(f"Could not attach ArcpyLogHandler: {e}")

    # --- Argument & Path Resolution ---
    if not args.optimize_script_path:
        args.optimize_script_path = _get_optimize_script_path(args.arc_mode or False)
    
    if not args.output_path:
        assert args.input_path is not None, "input_path must be set"
        # Handle both single file and directory for output path generation
        if isinstance(args.input_path, list):
            # If input is a list (from directory), use the directory path to create the output excel name
            # Assuming the list comes from a single directory, we can get the parent of the first file.
            if args.input_path:
                source_path = Path(args.input_path[0]).parent
                args.output_path = source_path / f"{source_path.name}_test.xlsx"
        else: # It's a single Path object
            source_path = Path(args.input_path)
            if source_path.is_dir():
                args.output_path = source_path / f"{source_path.name}_test.xlsx"
            else:
                args.output_path = source_path.with_name(f"{source_path.stem}_test.xlsx")

    if not args.temp_dir:
        args.temp_dir = DEFAULT_TEMP_DIR
    temp_directory = Path(args.temp_dir).resolve()
    temp_directory.mkdir(parents=True, exist_ok=True)
    logger.info(f"Using temporary directory: {temp_directory}")

    excel_writer = None
    try:
        logger.info("=== test_compression FUNCTION ENTRY ===")
        logger.info(f"Parsed args: input_path={args.input_path}, product_type={args.product_type}, arc_mode={args.arc_mode}")
        logger.info(f"Output excel: {args.output_path}, temp_dir: {args.temp_dir}")
        
        if args.csv_path:
            test_cases = _read_test_cases_from_csv(args.csv_path)
        elif args.product_type:
            test_cases = _load_compression_options(args.product_type)
        else:
            logger.error("No input method specified")
            return 1
            
        if not test_cases:
            logger.error("No test cases loaded. Exiting.")
            return 1

        if not args.output_path:
            logger.error("Output Excel path is required")
            return 1
        excel_writer = ExcelWriter(args.output_path)

        geotiff_files_str = get_geotiff_files(str(args.input_path))
        if not geotiff_files_str:
            logger.error(f"No GeoTIFF files found in {args.input_path}")
            return 1
        
        geotiff_files = [Path(p) for p in geotiff_files_str]

        for file_path in geotiff_files:
            # Mark the start of a new file group for border formatting
            excel_writer.mark_group_start()
            
            original_size_mb, original_compression_pct = _process_original_source_file(file_path, args, excel_writer)

            if original_compression_pct is None:
                logger.error(f"Failed to get compression percentage of the original source file {file_path}. Skipping.")
                continue

            script_generated_baseline_metrics = _process_script_baseline(file_path, args, excel_writer, test_cases[0][COL_TYPE], original_size_mb, original_compression_pct)
            
            if not script_generated_baseline_metrics.success_flag:
                logger.error(f"Failed to process baseline for {file_path}. Skipping.")
                continue

            _run_compression_tests(file_path, args, test_cases, excel_writer, script_generated_baseline_metrics, original_size_mb, original_compression_pct)
            
            # Mark the end of the file group and apply thick borders
            excel_writer.mark_group_end()

        excel_writer.save()
        logger.info("\nProcessing complete. Report saved.")

        if args.open_report and not args.arc_mode:
            try:
                logger.info(f"Attempting to open report: {args.output_path}")
                os.startfile(args.output_path)
                logger.info(f"Report opened: {args.output_path}")
            except Exception as e:
                logger.warning(f"Could not automatically open the report: {e}")

        logger.info("Exiting test_compression with success code 0.")
        return 0
    except Exception as e:
        logger.error(f"An unexpected error occurred in main: {e}")
        logger.error(traceback.format_exc())
        # Attempt to save any partial results on unexpected error
        if 'excel_writer' in locals() and excel_writer:
            excel_writer.save(success=False)
        return 1
    finally:
        shutdown_logger(logger)